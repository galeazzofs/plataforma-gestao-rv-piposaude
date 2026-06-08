"""Financial XLSX parser for the real "Consulta - Follow up Faturamento" format.

The spreadsheet has:
- Single sheet (any name; we take the first / active)
- Summary rows at top
- Header row somewhere in the first 20 rows (detected by "Cliente Mãe" / "Operadora")
- Data rows after the header

We parse, apply minimal filters (status RECEBIDO + period + non-empty),
and return rows ready to be persisted as financial_imports. The calculator
decides later which rows are MATCHED / UNMATCHED / PRODUTO_NAO_SUPORTADO / etc.
"""
import unicodedata
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook


class ParseError(Exception):
    pass


COLUMN_KEYWORDS = {
    'cliente_mae': lambda h: 'cliente' in h and 'mae' in h,
    'operadora': lambda h: 'operadora' in h,
    'produto': lambda h: 'produto' in h and 'segmenta' not in h,
    'numero_apolice': lambda h: 'apolice' in h and 'beneficio' not in h,
    'nf_valor_liquido': lambda h: 'nf' in h and 'liquido' in h,
    'data_recebimento': lambda h: 'data' in h and 'recebimento' in h,
    'mes_recebimento': lambda h: 'mes' in h and 'recebimento' in h,
    'status_recebimento': lambda h: 'status' in h and 'recebimento' in h,
    'tipo_receita': lambda h: 'tipo' in h and 'receita' in h,
}


def _normalize_header(s):
    if s is None:
        return ""
    decomp = unicodedata.normalize('NFD', str(s).lower().strip())
    return ''.join(c for c in decomp if unicodedata.category(c) != 'Mn')


def _detect_header_row(ws, scan_limit=20):
    """Find the header row by looking for 'cliente mae' in the first N rows."""
    max_cols = (ws.max_column or 40)
    for r in range(1, scan_limit + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                norm = _normalize_header(v)
                if 'cliente' in norm and 'mae' in norm:
                    return r
    raise ParseError(
        f"Header row not found in first {scan_limit} rows "
        "(looking for 'Cliente Mãe')"
    )


def _build_column_map(headers):
    """Map our field names → column index by scanning headers."""
    mapping = {}
    for col_idx, header in enumerate(headers, start=1):
        norm = _normalize_header(header)
        if not norm:
            continue
        for field, matcher in COLUMN_KEYWORDS.items():
            if field in mapping:
                continue
            if matcher(norm):
                mapping[field] = col_idx
    required = [
        'cliente_mae', 'operadora', 'produto',
        'nf_valor_liquido', 'data_recebimento', 'status_recebimento',
    ]
    missing = [f for f in required if f not in mapping]
    if missing:
        raise ParseError(f"Missing required columns: {missing}")
    return mapping


def _coerce_date(value):
    """Accept datetime, date, or 'dd/mm/yyyy' string. Returns date or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_financial_xlsx(filepath, target_year):
    """Parse the financial XLSX, returning rows that pass minimal filters.

    Each row keeps its own competência month (mes_recebimento, YYYY-MM); the
    persistence layer tags FinancialImport.month per row from it. A single
    upload of a multi-month export therefore populates every month.

    Filters applied in the parser:
      - status_recebimento != 'RECEBIDO' → drop
      - cliente_mae or nf_valor_liquido empty → drop
      - data_recebimento year != target_year → drop

    Rows with unsupported products (Mental, Fitness) and Odonto/Vida/Saúde
    all PASS the parser. The calculator decides later.

    Returns:
        {
          'rows': [{cliente_mae, operadora, produto, nf_valor_liquido,
                    data_recebimento, mes_recebimento, tipo_receita,
                    status_recebimento, _row}],
          'stats': {total_lidas, descartadas_status, descartadas_periodo,
                    descartadas_vazias, persistidas}
        }

    Raises:
        ParseError: header row cannot be found, or required columns missing.
        FileNotFoundError: filepath does not exist.
    """
    wb = load_workbook(filepath, read_only=False, data_only=True)
    ws = wb.active

    header_row = _detect_header_row(ws)
    max_cols = (ws.max_column or 40)
    headers = [ws.cell(row=header_row, column=c).value
               for c in range(1, max_cols + 1)]
    cmap = _build_column_map(headers)

    rows = []
    stats = {
        'total_lidas': 0,
        'descartadas_status': 0,
        'descartadas_periodo': 0,
        'descartadas_vazias': 0,
        'descartadas_tipo_receita': 0,
        'persistidas': 0,
    }

    def cell(r, field):
        idx = cmap.get(field)
        return ws.cell(row=r, column=idx).value if idx else None

    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        cliente_mae = cell(r, 'cliente_mae')
        nf_liq = cell(r, 'nf_valor_liquido')

        # Count "read" only if the row has any content at all
        if cliente_mae is None and nf_liq is None:
            continue
        stats['total_lidas'] += 1

        if not cliente_mae or nf_liq is None:
            stats['descartadas_vazias'] += 1
            continue

        status = cell(r, 'status_recebimento')
        if (status or '').strip().upper() != 'RECEBIDO':
            stats['descartadas_status'] += 1
            continue

        data_rec = _coerce_date(cell(r, 'data_recebimento'))
        if data_rec is None:
            stats['descartadas_vazias'] += 1
            continue

        if data_rec.year != target_year:
            stats['descartadas_periodo'] += 1
            continue

        # Only keep Comissão and Agenciamento revenue types (skip Premiação etc.)
        tipo = _normalize_header(cell(r, 'tipo_receita') or '')
        if tipo and tipo not in ('comissao', 'agenciamento'):
            stats['descartadas_tipo_receita'] += 1
            continue

        # ALWAYS derive YYYY-MM from data_recebimento. The XLSX's
        # "Mês Recebimento" column is unreliable (sometimes a full English
        # month name like "September" which doesn't fit varchar(7), sometimes
        # YYYY-MM, sometimes a date). Trust data_recebimento instead.
        mes_rec = data_rec.strftime("%Y-%m")

        # NF values are money — Decimal preserves cents, float doesn't.
        try:
            valor_liquido = Decimal(str(nf_liq))
        except (InvalidOperation, ValueError, TypeError):
            stats['descartadas_vazias'] += 1
            continue

        numero_raw = cell(r, 'numero_apolice')
        numero_apolice = str(numero_raw).strip() if numero_raw is not None else ''

        rows.append({
            'cliente_mae': str(cliente_mae).strip(),
            'operadora': str(cell(r, 'operadora') or '').strip(),
            'produto': str(cell(r, 'produto') or '').strip(),
            'numero_apolice': numero_apolice or None,
            'nf_valor_liquido': valor_liquido,
            'data_recebimento': data_rec,
            'mes_recebimento': mes_rec,
            'tipo_receita': (str(cell(r, 'tipo_receita') or '').strip() or None),
            'status_recebimento': 'RECEBIDO',
            '_row': r,
        })
        stats['persistidas'] += 1

    wb.close()
    return {'rows': rows, 'stats': stats}
