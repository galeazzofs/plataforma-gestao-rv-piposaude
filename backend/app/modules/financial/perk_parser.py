"""Perk/subsidy parser (XLSX or CSV).

Parses spreadsheets with columns like:
  - Cliente Pipo / Cliente
  - Valor
  - Mês (Competência)
  - Ano

Returns rows ready to be persisted as Perk records. Handles both US
("3,934.02", the Omie export) and BR ("3.500,00") number formats.
"""
import csv as _csv
import unicodedata
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


class PerkParseError(Exception):
    pass


def parse_money(raw):
    """Parse a money value in either US (``3,934.02``) or BR (``3.500,00``)
    format into a signed Decimal.

    Returns ``None`` when the value is blank/None or has no digits. Parentheses
    mean negative. Numeric inputs (int/float — e.g. an xlsx numeric cell) pass
    straight through without separator guessing.

    Format detection: when both ``,`` and ``.`` are present, the RIGHTMOST is
    the decimal mark and the other is thousands grouping. With a single
    separator, exactly two trailing digits is treated as a decimal mark;
    anything else (multiple occurrences, or not two trailing digits) is
    thousands grouping and stripped.
    """
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))

    s = str(raw).strip()
    if not s:
        return None

    negative = False
    if s.startswith('(') and s.endswith(')'):
        s = s[1:-1].strip()
        negative = True
    if s.startswith('-'):
        negative = True
        s = s[1:].strip()

    if not any(ch.isdigit() for ch in s):
        return None

    has_comma = ',' in s
    has_dot = '.' in s
    if has_comma and has_dot:
        if s.rfind(',') > s.rfind('.'):
            decimal_sep, grouping_sep = ',', '.'
        else:
            decimal_sep, grouping_sep = '.', ','
        s = s.replace(grouping_sep, '').replace(decimal_sep, '.')
    elif has_comma or has_dot:
        sep = ',' if has_comma else '.'
        trailing = len(s) - s.rfind(sep) - 1
        if s.count(sep) == 1 and trailing == 2:
            s = s.replace(sep, '.')
        else:
            s = s.replace(sep, '')

    try:
        value = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return -value if negative else value


COLUMN_KEYWORDS = {
    'cliente': lambda h: 'cliente' in h,
    'valor': lambda h: h == 'valor' or ('valor' in h and 'liquido' not in h),
    'mes': lambda h: 'mes' in h or 'competencia' in h,
    'ano': lambda h: h == 'ano',
}


def _normalize_header(s):
    if s is None:
        return ""
    decomp = unicodedata.normalize('NFD', str(s).lower().strip())
    return ''.join(c for c in decomp if unicodedata.category(c) != 'Mn')


def _row_is_header(cells):
    """A real header row has BOTH a 'cliente' and a 'valor' cell — title rows
    that only mention 'Cliente' shouldn't be picked up."""
    norms = [_normalize_header(v) for v in cells if v is not None]
    has_cliente = any('cliente' in n for n in norms)
    has_valor = any(
        n == 'valor' or ('valor' in n and 'liquido' not in n) for n in norms
    )
    return has_cliente and has_valor


def _detect_header_row(grid, scan_limit=20):
    """Return the 0-based index of the header row within the grid."""
    for r in range(min(scan_limit, len(grid))):
        if _row_is_header(grid[r]):
            return r
    raise PerkParseError(
        "Header row not found — expected a row with both 'Cliente' and "
        "'Valor' columns in the first 20 rows."
    )


def _build_column_map(headers):
    """Map each field to its 0-based column index from a header row."""
    cmap = {}
    for col_idx, raw in enumerate(headers):
        norm = _normalize_header(raw)
        if not norm:
            continue
        for field, matcher in COLUMN_KEYWORDS.items():
            if field not in cmap and matcher(norm):
                cmap[field] = col_idx
    if 'cliente' not in cmap:
        raise PerkParseError("Column 'Cliente' not found in header row.")
    if 'valor' not in cmap:
        raise PerkParseError("Column 'Valor' not found in header row.")
    return cmap


def _parse_month(raw):
    """Extract month number from values like '01 - Janeiro', '3', 'Março', etc."""
    if raw is None:
        return None
    s = str(raw).strip()
    # "01 - Janeiro" → take first part
    if '-' in s:
        s = s.split('-')[0].strip()
    try:
        return int(s)
    except ValueError:
        return None


def _is_blank(v):
    return v is None or str(v).strip() == ''


def _load_xlsx_grid(filepath):
    """Read an XLSX into a list-of-rows grid (read_only streams cells)."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    max_cols = min(ws.max_column or 40, 40)
    grid = [
        list(row)
        for row in ws.iter_rows(min_col=1, max_col=max_cols, values_only=True)
    ]
    wb.close()
    return grid


def _load_csv_grid(filepath):
    """Read a CSV into a list-of-rows grid. utf-8-sig tolerates a BOM."""
    with open(filepath, newline='', encoding='utf-8-sig') as fh:
        return [list(row) for row in _csv.reader(fh)]


def _parse_perk_grid(grid, target_year):
    """Core parser shared by the XLSX and CSV front-ends.

    Returns:
        {
          'rows': [{client_name, amount, month, year}],
          'stats': {total_lidas, descartadas_periodo, descartadas_vazias, persistidas}
        }
    """
    header_idx = _detect_header_row(grid)
    cmap = _build_column_map(grid[header_idx])

    rows = []
    stats = {
        'total_lidas': 0,
        'descartadas_periodo': 0,
        'descartadas_vazias': 0,
        'persistidas': 0,
    }

    def cell(row, field):
        idx = cmap.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for r in range(header_idx + 1, len(grid)):
        row = grid[r]
        cliente = cell(row, 'cliente')
        valor_raw = cell(row, 'valor')

        if _is_blank(cliente) and _is_blank(valor_raw):
            continue
        stats['total_lidas'] += 1

        if _is_blank(cliente) or _is_blank(valor_raw):
            stats['descartadas_vazias'] += 1
            continue

        # Perks are stored as positive magnitudes; the sheet carries them as
        # parenthesized costs ("(3,934.02)"). parse_money reads US or BR format.
        amount = parse_money(valor_raw)
        if amount is None:
            stats['descartadas_vazias'] += 1
            continue
        amount = abs(amount)
        if amount <= 0:
            stats['descartadas_vazias'] += 1
            continue

        # Parse period
        ano_raw = cell(row, 'ano')
        mes_raw = cell(row, 'mes')

        try:
            year = int(ano_raw) if not _is_blank(ano_raw) else target_year
        except (ValueError, TypeError):
            year = target_year

        month = _parse_month(mes_raw)
        if month is None or month < 1 or month > 12:
            stats['descartadas_vazias'] += 1
            continue

        if year != target_year:
            stats['descartadas_periodo'] += 1
            continue

        rows.append({
            'client_name': str(cliente).strip(),
            'amount': amount,
            'month': month,
            'year': year,
            '_row': r + 1,
        })
        stats['persistidas'] += 1

    return {'rows': rows, 'stats': stats}


def parse_perk_xlsx(filepath, target_year):
    """Parse a perk XLSX for `target_year` (each row keeps its own competência
    month from the 'Mês' column)."""
    return _parse_perk_grid(_load_xlsx_grid(filepath), target_year)


def parse_perk_csv(filepath, target_year):
    """Parse a perk CSV (UTF-8) for `target_year`, keeping each row's own
    competência month."""
    return _parse_perk_grid(_load_csv_grid(filepath), target_year)


def parse_perk_file(filepath, target_year, original_filename=None):
    """Parse a perk sheet, dispatching to CSV or XLSX by extension.

    `original_filename` (the uploaded name) takes precedence over `filepath`
    because the upload route saves to a temp file whose suffix may not reflect
    the real format.
    """
    name = (original_filename or filepath or '').lower()
    if name.endswith('.csv'):
        return parse_perk_csv(filepath, target_year)
    return parse_perk_xlsx(filepath, target_year)
